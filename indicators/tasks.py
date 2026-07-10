from celery import shared_task
from datetime import datetime, timedelta, timezone
from django.db.models import Sum, Avg, F, FloatField, Max, Count, Min, Q, QuerySet, Value
from django.db.models.functions import Cast, TruncDay, Greatest
import logging
import calendar
from django.utils import timezone as django_timezone
import pytz
from collections import defaultdict
import statistics
import os
import tempfile
import csv

from scada_proxy.models import (
    Device, Institution, DeviceCategory, TaskProgress,
    MeterMeasurement, InverterMeasurement, WeatherStationMeasurement,
)
from .energy import consumption_energy_kwh, generation_energy_kwh, SAMPLE_INTERVAL_HOURS
from .models import (
    MonthlyConsumptionKPI,
    DailyChartData,
    InverterIndicators,
    InverterChartData,
    WeatherStationIndicators,
    WeatherStationChartData,
    GeneratedReport
)

from core.task_locks import single_instance

logger = logging.getLogger(__name__)

# Helpers puros extraídos a services/ (Ola 5); re-exportados para no romper los imports
# existentes (`from indicators.tasks import ...`).
from indicators.services.sanitize import (  # noqa: E402,F401
    ROLLOVER_CAP_FACTOR,
    ROLLOVER_CAP_MARGIN_KWH,
    _accumulate_register_energy,
)
from indicators.services.date_ranges import (  # noqa: E402,F401
    COLOMBIA_TZ,
    colombia_day_range,
    get_colombia_date,
    get_colombia_now,
)
from indicators.services.device_calc import run_over_days, run_over_months  # noqa: E402
from indicators.services.meter_calc import METER_FIELDS, compute_meter_indicators  # noqa: E402
from indicators.services.inverter_calc import INVERTER_FIELDS, compute_inverter_indicators  # noqa: E402
from indicators.services.rows import _row_get  # noqa: E402,F401

@shared_task(bind=True, retry_backoff=60, max_retries=3)
@single_instance('calculate-monthly-kpi', ttl=1200)
def calculate_monthly_consumption_kpi(self):
    """
    Calcula el consumo total, la generación total, la potencia instantánea promedio,
    la temperatura promedio diaria, la humedad relativa promedio, y la velocidad del viento promedio
    para el mes actual (hasta la fecha) y el mes anterior (hasta el mismo día).
    Guarda los resultados en MonthlyConsumptionKPI.
    """
    logger.info("=== INICIANDO TAREA: calculate_monthly_consumption_kpi ===")
    try:
        # Obtener la fecha y hora actual en zona horaria de Colombia
        today = get_colombia_date()
        current_day = today.day

        logger.info(f"Fecha actual en Colombia: {today}, día del mes: {current_day}")

        # --- Cálculo de rangos de fechas para el mes actual ---
        # El mes actual va desde el primer día hasta la fecha actual
        start_current_month = today.replace(day=1)
        end_current_month = today

        # --- Lógica de cálculo de rangos de fechas para el mes anterior ---
        first_day_current_month = today.replace(day=1)
        last_day_previous_month = first_day_current_month - timedelta(days=1)
        previous_month = last_day_previous_month.month
        previous_year = last_day_previous_month.year

        start_previous_month = last_day_previous_month.replace(day=1)
        
        # El final del rango del mes anterior es el mismo día que el mes actual,
        # pero ajustado por si el mes anterior tiene menos días (ej. febrero).
        last_day_of_previous_month = calendar.monthrange(previous_year, previous_month)[1]
        day_for_previous_month = min(current_day, last_day_of_previous_month)
        end_previous_month = last_day_previous_month.replace(day=day_for_previous_month)

        logger.info(f"Rango mes actual: {start_current_month} -> {end_current_month}")
        logger.info(f"Rango mes anterior: {start_previous_month} -> {end_previous_month}")

        # Obtener los dispositivos activos de cada categoría
        electric_meters = Device.objects.filter(category__name__iexact='electricMeter', is_active=True)
        inverters = Device.objects.filter(category__name__iexact='inverter', is_active=True)
        weather_stations = Device.objects.filter(category__name__iexact='weatherStation', is_active=True) 
        
        logger.info(f"Dispositivos encontrados:")
        logger.info(f"  - Medidores eléctricos: {electric_meters.count()} dispositivos")
        logger.info(f"  - Inversores: {inverters.count()} dispositivos")
        logger.info(f"  - Estaciones meteorológicas: {weather_stations.count()} dispositivos")

        # --- Cálculo de Consumo Total (Medidores Eléctricos) ---
        logger.info("Calculando consumo total (medidores eléctricos)...")
        # Cambiar el cálculo de consumo para que sea consistente
        # Energía consumida = Σ(totalActivePower) · Δt (ver indicators/energy.py).
        # totalActivePower está en kW (confirmado por scripts/audit_indicators.py:
        # mediana ~0.9, máx ~74), así que NO se divide por 1000.
        # ANTES: se guardaba la SUMA cruda de la potencia como si fuera kWh (sin integrar
        # Δt y con el mensual incoherente con el diario).
        # Consumo NETO = Σ(totalActivePower) incluyendo negativos (inyección a la red,
        # net metering). Consumo BRUTO = Σ(max(totalActivePower, 0)), solo energía tomada
        # de la red (los periodos de exportación cuentan como 0). Ambos en kW → factor 1.
        # `Greatest(x, 0)` clampa los negativos a nivel de agregación en la BD.
        # Rangos aware [inicio 00:00, fin+1día 00:00) en Bogotá: mismos límites
        # que el antiguo date__date__range inclusivo, pero como comparación de
        # timestamps que usa el índice UNIQUE (device, date). Lecturas sobre el
        # esquema v2 (columnas tipadas): sin Cast ni parseo de jsonb.
        cur_a, cur_b = colombia_day_range(start_current_month, end_current_month)
        prev_a, prev_b = colombia_day_range(start_previous_month, end_previous_month)

        def _meter_window(a, b):
            return MeterMeasurement.objects.filter(
                device__in=electric_meters,
                date__gte=a, date__lt=b,
                totalActivePower__isnull=False,
            ).aggregate(
                net_sum=Sum('totalActivePower'),
                gross_sum=Sum(Greatest(F('totalActivePower'), Value(0.0))),
            )

        current_agg = _meter_window(cur_a, cur_b)
        current_month_consumption_sum = consumption_energy_kwh(current_agg['net_sum'])
        current_month_gross_consumption_sum = consumption_energy_kwh(current_agg['gross_sum'])

        previous_agg = _meter_window(prev_a, prev_b)
        previous_month_consumption_sum = consumption_energy_kwh(previous_agg['net_sum'])
        previous_month_gross_consumption_sum = consumption_energy_kwh(previous_agg['gross_sum'])
        logger.info(f"Consumo NETO - Mes actual: {current_month_consumption_sum:.2f} kWh, Mes anterior: {previous_month_consumption_sum:.2f} kWh")
        logger.info(f"Consumo BRUTO - Mes actual: {current_month_gross_consumption_sum:.2f} kWh, Mes anterior: {previous_month_gross_consumption_sum:.2f} kWh")

        # --- Cálculo de Generación Total (Inversores) ---
        logger.info("Calculando generación total (inversores)...")
        # Energía generada = Σ(acPower de TODOS los inversores) · Δt / 1000, misma fórmula
        # canónica que el consumo (ver indicators/energy.py).
        # acPower está en Watts (confirmado por auditoría: mediana ~2936, máx ~16266),
        # por eso aquí SÍ se divide por 1000 (a diferencia del consumo).
        # ANTES: se agrupaba por día y se calculaba (Σ acPower / nº mediciones de la flota) · 24,
        # es decir, la potencia media POR MUESTRA de la flota × 24 h → energía de UN inversor
        # "promedio", no la suma de los N inversores. Eso dividía la generación total por N.
        def _inverter_window(a, b):
            # Suma (para energía) y promedio (potencia instantánea) de acPower
            # en UNA sola query por ventana.
            return InverterMeasurement.objects.filter(
                device__in=inverters,
                date__gte=a, date__lt=b,
                acPower__isnull=False,
            ).aggregate(total_sum=Sum('acPower'), avg_value=Avg('acPower'))

        inv_current = _inverter_window(cur_a, cur_b)
        inv_previous = _inverter_window(prev_a, prev_b)

        current_month_generation_sum = generation_energy_kwh(inv_current['total_sum'])
        previous_month_generation_sum = generation_energy_kwh(inv_previous['total_sum'])
        logger.info(f"Generación total - Mes actual: {current_month_generation_sum:.2f} kWh, Mes anterior: {previous_month_generation_sum:.2f} kWh")

        # --- Cálculo de Potencia Instantánea Promedio (Inversores) ---
        avg_instantaneous_power_current = inv_current['avg_value'] or 0.0
        avg_instantaneous_power_previous = inv_previous['avg_value'] or 0.0
        logger.info(f"Potencia instantánea promedio - Mes actual: {avg_instantaneous_power_current:.2f} W, Mes anterior: {avg_instantaneous_power_previous:.2f} W")

        # --- Métricas meteorológicas (temperatura, humedad, viento, irradiancia) ---
        # Una sola query por ventana: Avg de SQL ignora NULL por métrica, así
        # que el resultado es idéntico a los antiguos filtros isnull separados.
        logger.info("Calculando promedios meteorológicos (estaciones)...")

        def _weather_window(a, b):
            return WeatherStationMeasurement.objects.filter(
                device__in=weather_stations,
                date__gte=a, date__lt=b,
            ).aggregate(
                temp=Avg('temperature'),
                humidity=Avg('humidity'),
                wind=Avg('windSpeed'),
                irradiance=Avg('irradiance'),
            )

        wx_current = _weather_window(cur_a, cur_b)
        wx_previous = _weather_window(prev_a, prev_b)

        avg_daily_temp_current = wx_current['temp'] or 0.0
        avg_daily_temp_previous = wx_previous['temp'] or 0.0
        logger.info(f"Temperatura promedio diaria - Mes actual: {avg_daily_temp_current:.2f} °C, Mes anterior: {avg_daily_temp_previous:.2f} °C")

        avg_relative_humidity_current = wx_current['humidity'] or 0.0
        avg_relative_humidity_previous = wx_previous['humidity'] or 0.0
        logger.info(f"Humedad relativa promedio - Mes actual: {avg_relative_humidity_current:.2f} %RH, Mes anterior: {avg_relative_humidity_previous:.2f} %RH")

        avg_wind_speed_current = wx_current['wind'] or 0.0
        avg_wind_speed_previous = wx_previous['wind'] or 0.0
        logger.info(f"Velocidad del viento promedio - Mes actual: {avg_wind_speed_current:.2f} km/h, Mes anterior: {avg_wind_speed_previous:.2f} km/h")

        avg_irradiance_current = wx_current['irradiance'] or 0.0
        avg_irradiance_previous = wx_previous['irradiance'] or 0.0
        logger.info(f"Irradiancia solar promedio - Mes actual: {avg_irradiance_current:.2f} W/m², Mes anterior: {avg_irradiance_previous:.2f} W/m²")

        # Guardar en la base de datos
        logger.info("Guardando KPIs mensuales en la base de datos...")
        MonthlyConsumptionKPI.objects.update_or_create(
            pk=1,
            defaults={
                'total_consumption_current_month': current_month_consumption_sum,
                'total_consumption_previous_month': previous_month_consumption_sum,
                'total_gross_consumption_current_month': current_month_gross_consumption_sum,
                'total_gross_consumption_previous_month': previous_month_gross_consumption_sum,
                'total_generation_current_month': current_month_generation_sum,
                'total_generation_previous_month': previous_month_generation_sum,
                'avg_instantaneous_power_current_month': avg_instantaneous_power_current,
                'avg_instantaneous_power_previous_month': avg_instantaneous_power_previous,
                'avg_daily_temp_current_month': avg_daily_temp_current,
                'avg_daily_temp_previous_month': avg_daily_temp_previous,
                'avg_relative_humidity_current_month': avg_relative_humidity_current,
                'avg_relative_humidity_previous_month': avg_relative_humidity_previous,
                'avg_wind_speed_current_month': avg_wind_speed_current,
                'avg_wind_speed_previous_month': avg_wind_speed_previous,
                'avg_irradiance_current_month': avg_irradiance_current,
                'avg_irradiance_previous_month': avg_irradiance_previous,
            }
        )

        logger.info("=== TAREA COMPLETADA: calculate_monthly_consumption_kpi ===")
        logger.info("Todos los KPIs mensuales han sido calculados y actualizados exitosamente.")
        return "All Monthly KPIs calculated and updated successfully."

    except Exception as e:
        logger.error(f"=== ERROR EN TAREA: calculate_monthly_consumption_kpi ===")
        logger.error(f"Error calculando KPIs mensuales: {e}", exc_info=True)
        raise

@shared_task(bind=True, retry_backoff=30, max_retries=3)
@single_instance('calculate-daily-chart', ttl=1200)
def calculate_and_save_daily_data(self, start_date_str: str = None, end_date_str: str = None):
    """
    Calcula el consumo, la generación, el balance de energía y la temperatura promedio diaria 
    para un rango de fechas.
    """
    logger.info("=== INICIANDO TAREA: calculate_and_save_daily_data ===")
    try:
        # Si no se proporcionan fechas, calcular para el día anterior en Colombia
        if not start_date_str or not end_date_str:
            yesterday = get_colombia_date() - timedelta(days=1)
            start_date_str = yesterday.isoformat()
            end_date_str = yesterday.isoformat()
            logger.info(f"No se proporcionaron fechas, calculando para el día anterior en Colombia: {yesterday}")

        # 1. Convertir strings a objetos datetime conscientes de la zona horaria de Colombia
        start_date = datetime.fromisoformat(start_date_str).replace(tzinfo=COLOMBIA_TZ)
        end_date = datetime.fromisoformat(end_date_str).replace(tzinfo=COLOMBIA_TZ)

        logger.info(f"Rango de fechas a procesar en Colombia: {start_date.date()} -> {end_date.date()}")
        
        # 2. Obtener los dispositivos eléctricos, inversores y estaciones meteorológicas
        electric_meters: QuerySet[Device] = Device.objects.filter(category__name__iexact='electricMeter', is_active=True)
        inverters: QuerySet[Device] = Device.objects.filter(category__name__iexact='inverter', is_active=True)
        weather_stations: QuerySet[Device] = Device.objects.filter(category__name__iexact='weatherStation', is_active=True)
        
        logger.info(f"Dispositivos encontrados para cálculo diario:")
        logger.info(f"  - Medidores eléctricos: {electric_meters.count()} dispositivos")
        logger.info(f"  - Inversores: {inverters.count()} dispositivos")
        logger.info(f"  - Estaciones meteorológicas: {weather_stations.count()} dispositivos")
        
        if not electric_meters.exists() and not inverters.exists() and not weather_stations.exists():
            logger.warning("No se encontraron dispositivos activos para procesar.")
            return

        electric_meter_ids = electric_meters.values_list('id', flat=True)
        inverter_ids = inverters.values_list('id', flat=True)
        weather_station_ids = weather_stations.values_list('id', flat=True)

        # 3. Iterar día por día en el rango especificado
        current_date = start_date
        total_days_processed = 0
        total_records_created = 0
        total_records_updated = 0
        
        logger.info(f"Iniciando procesamiento de {((end_date - start_date).days + 1)} días...")
        
        # v2: 3 queries agrupadas por día de Bogotá sobre TODO el rango (antes:
        # 4 queries por CADA día). TruncDay(tzinfo=COLOMBIA_TZ) reproduce el
        # corte del antiguo `date__date=<día>` con TIME_ZONE=America/Bogota.
        range_a, range_b = colombia_day_range(start_date.date(), end_date.date())
        day_trunc = TruncDay('date', tzinfo=COLOMBIA_TZ)

        meter_by_day = {
            r['d'].date(): r for r in MeterMeasurement.objects.filter(
                device__in=electric_meter_ids,
                date__gte=range_a, date__lt=range_b,
                totalActivePower__isnull=False,
            ).annotate(d=day_trunc).values('d').annotate(
                net=Sum('totalActivePower'),
                gross=Sum(Greatest(F('totalActivePower'), Value(0.0))),
            )
        }
        inverter_by_day = {
            r['d'].date(): r for r in InverterMeasurement.objects.filter(
                device__in=inverter_ids,
                date__gte=range_a, date__lt=range_b,
                acPower__isnull=False,
            ).annotate(d=day_trunc).values('d').annotate(gen=Sum('acPower'))
        }
        weather_by_day = {
            r['d'].date(): r for r in WeatherStationMeasurement.objects.filter(
                device__in=weather_station_ids,
                date__gte=range_a, date__lt=range_b,
            ).annotate(d=day_trunc).values('d').annotate(
                temp=Avg('temperature'), wind=Avg('windSpeed'), irr=Avg('irradiance'),
            )
        }

        while current_date <= end_date:
            single_date = current_date.date()
            logger.info(f"Procesando fecha en Colombia: {single_date}")

            m_day = meter_by_day.get(single_date, {})
            i_day = inverter_by_day.get(single_date, {})
            w_day = weather_by_day.get(single_date, {})

            daily_consumption_sum = m_day.get('net')
            daily_gross_consumption_sum = m_day.get('gross')
            daily_generation_sum = i_day.get('gen')
            daily_temp_avg = w_day.get('temp') or 0.0
            avg_wind_speed = w_day.get('wind') or 0.0
            avg_irradiance = w_day.get('irr') or 0.0

            # Consumo y generación con la fórmula canónica Σ(P)·Δt (ver indicators/energy.py),
            # idéntica a la del KPI mensual. La unidad de potencia difiere por métrica:
            # consumo (totalActivePower) en kW; generación (acPower) en W (÷1000).
            # ANTES: el consumo solo dividía /1000 (sin integrar Δt) y la generación
            # promediaba sobre la flota (avg_power_w · 24), dividiendo por el nº de inversores.
            daily_consumption_kwh = consumption_energy_kwh(daily_consumption_sum)
            daily_gross_consumption_kwh = consumption_energy_kwh(daily_gross_consumption_sum)
            daily_generation_kwh = generation_energy_kwh(daily_generation_sum)

            # Calcular balance energético (ambos en kWh)
            daily_balance_sum = daily_generation_kwh - daily_consumption_kwh

            daily_data_obj, created = DailyChartData.objects.update_or_create(
                date=single_date,
                defaults={
                    'daily_consumption': daily_consumption_kwh,  # NETO, en kWh
                    'daily_gross_consumption': daily_gross_consumption_kwh,  # BRUTO, en kWh
                    'daily_generation': daily_generation_kwh,    # Ahora en kWh
                    'daily_balance': daily_balance_sum,
                    'avg_daily_temp': daily_temp_avg,
                    'avg_wind_speed': avg_wind_speed,
                    'avg_irradiance': avg_irradiance
                }
            )
            
            if created:
                total_records_created += 1
                action = "creado"
            else:
                total_records_updated += 1
                action = "actualizado"
                
            logger.info(f"  Dato diario {action} para {single_date}:")
            logger.info(f"    - Consumo: {daily_consumption_kwh:.2f} kWh")
            logger.info(f"    - Generación: {daily_generation_kwh:.2f} kWh")
            logger.info(f"    - Balance: {daily_balance_sum:.2f} kWh")
            logger.info(f"    - Temperatura promedio: {daily_temp_avg:.2f} °C")
            logger.info(f"    - Velocidad del viento promedio: {avg_wind_speed:.2f} km/h")
            logger.info(f"    - Irradiancia promedio: {avg_irradiance:.2f} W/m²")

            current_date += timedelta(days=1)
            total_days_processed += 1

        logger.info("=== RESUMEN DE PROCESAMIENTO ===")
        logger.info(f"Días procesados: {total_days_processed}")
        logger.info(f"Registros creados: {total_records_created}")
        logger.info(f"Registros actualizados: {total_records_updated}")
        logger.info("=== TAREA COMPLETADA: calculate_and_save_daily_data ===")

    except Exception as e:
        logger.error(f"=== ERROR EN TAREA: calculate_and_save_daily_data ===")
        logger.error(f"Error en el cálculo de datos diarios: {e}", exc_info=True)
        raise

def _calculate_hourly_consumption(measurements):
    """
    Calcula el consumo por hora del día basado en las mediciones.

    Recibe filas dict (v2: `.values('date', 'totalActivePower')`).
    """
    hourly_data = defaultdict(list)

    for row in measurements:
        hour = row['date'].hour
        power_value = _row_get(row, 'totalActivePower')
        hourly_data[hour].append(power_value)
    
    # Calcular promedio por hora
    hourly_consumption = [0.0] * 24
    for hour in range(24):
        if hour in hourly_data:
            hourly_consumption[hour] = sum(hourly_data[hour]) / len(hourly_data[hour])
    
    return hourly_consumption

# indicators/tasks.py
@shared_task
def calculate_electric_meter_indicators(device_id, date_str, time_range='daily'):
    """
    Calcula todos los indicadores eléctricos para un medidor específico en una fecha dada.
    """
    try:
        from datetime import datetime, timedelta
        from django.db.models import Max, Min, Avg
        from scada_proxy.models import Device, Institution
        from .models import ElectricMeterIndicators
        
        # Parsear la fecha
        if isinstance(date_str, str):
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            date = date_str
        
        # Obtener el dispositivo y la institución
        device = Device.objects.get(id=device_id)
        institution = device.institution
        
        # Determinar el rango de fechas
        if time_range == 'daily':
            start_date = date
            end_date = date + timedelta(days=1)
        else:  # monthly
            start_date = date.replace(day=1)
            if date.month == 12:
                end_date = date.replace(year=date.year + 1, month=1, day=1)
            else:
                end_date = date.replace(month=date.month + 1, day=1)
        
        # Obtener todas las mediciones del período (v2: rango aware Bogotá,
        # equivalente al antiguo date__gte/lt con objetos date interpretados
        # a medianoche en TIME_ZONE=America/Bogota)
        start_dt, end_dt = colombia_day_range(start_date, end_date - timedelta(days=1))
        measurements = MeterMeasurement.objects.filter(
            device=device,
            date__gte=start_dt,
            date__lt=end_dt
        ).order_by('date')

        if not measurements.exists():
            return f"No hay mediciones para {device.name} en {date}"

        # Cálculo puro extraído a services/meter_calc.py (Ola 5).
        computed = compute_meter_indicators(
            measurements.values(*METER_FIELDS).iterator(chunk_size=2000)
        )

        # Guardar o actualizar los indicadores
        indicators, created = ElectricMeterIndicators.objects.update_or_create(
            device=device,
            institution=institution,
            date=date,
            time_range=time_range,
            defaults={
                **computed,
                'measurement_count': measurements.count(),
                'last_measurement_date': measurements.last().date if measurements.exists() else None,
            }
        )

        action = "creado" if created else "actualizado"
        return f"Indicadores eléctricos {action} para {device.name} en {date} ({time_range})"
        
    except Exception as e:
        return f"Error calculando indicadores eléctricos: {str(e)}"


@shared_task
def calculate_inverter_indicators(device_id, date_str, time_range='daily'):
    """
    Calcula todos los indicadores de inversores para un dispositivo específico en una fecha dada.
    """
    try:
        from datetime import datetime, timedelta
        from django.db.models import Max, Min, Avg, StdDev
        from scada_proxy.models import Device, Institution
        from .models import InverterIndicators, InverterChartData
        
        # Parsear la fecha
        if isinstance(date_str, str):
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            date = date_str
        
        # Obtener el dispositivo y la institución
        device = Device.objects.get(id=device_id)
        institution = device.institution
        
        # Determinar el rango de fechas
        if time_range == 'daily':
            start_date = date
            end_date = date + timedelta(days=1)
        else:  # monthly
            start_date = date.replace(day=1)
            if date.month == 12:
                end_date = date.replace(year=date.year + 1, month=1, day=1)
            else:
                end_date = date.replace(month=date.month + 1, day=1)
        
        # Obtener todas las mediciones del período (v2: rango aware Bogotá,
        # equivalente al antiguo date__gte/lt con objetos date interpretados
        # a medianoche en TIME_ZONE=America/Bogota)
        start_dt, end_dt = colombia_day_range(start_date, end_date - timedelta(days=1))
        measurements = InverterMeasurement.objects.filter(
            device=device,
            date__gte=start_dt,
            date__lt=end_dt
        ).order_by('date')

        if not measurements.exists():
            return f"No hay mediciones para {device.name} en {date}"
        
        # Cálculo puro extraído a services/inverter_calc.py (Ola 5).
        computed = compute_inverter_indicators(
            measurements.values(*INVERTER_FIELDS).iterator(chunk_size=2000)
        )

        # Guardar o actualizar los indicadores
        indicators, created = InverterIndicators.objects.update_or_create(
            device=device,
            institution=institution,
            date=date,
            time_range=time_range,
            defaults={
                **computed,
                'measurement_count': measurements.count(),
                'last_measurement_date': measurements.last().date if measurements.exists() else None,
            }
        )

        # Crear datos para gráficos
        hourly_data = _calculate_hourly_inverter_data(
            measurements.values('date', 'acPower', 'dcPower').iterator(chunk_size=2000)
        )
        
        chart_data, chart_created = InverterChartData.objects.update_or_create(
            device=device,
            institution=institution,
            date=date,
            defaults=hourly_data
        )
        
        action = "creado" if created else "actualizado"
        return f"Indicadores de inversor {action} para {device.name} en {date} ({time_range})"
        
    except Exception as e:
        return f"Error calculando indicadores de inversor: {str(e)}"


def _calculate_hourly_inverter_data(measurements):
    """
    Calcula datos por hora para gráficos de inversores.

    Recibe filas dict (v2: `.values('date', 'acPower', 'dcPower')`).
    """
    hourly_efficiency = [0] * 24
    hourly_generation = [0] * 24
    hourly_irradiance = [0] * 24
    hourly_temperature = [0] * 24
    hourly_dc_power = [0] * 24
    hourly_ac_power = [0] * 24

    hourly_counts = [0] * 24

    for row in measurements:
        hour = row['date'].hour

        # Acumular valores por hora
        if row['acPower'] is not None:
            hourly_ac_power[hour] += row['acPower']
            hourly_counts[hour] += 1

        if row['dcPower'] is not None:
            hourly_dc_power[hour] += row['dcPower']

        # v2: la categoría 'inverter' no tiene columnas irradiance/temperature
        # (en v1 data.get(...) devolvía None y nunca se acumulaba nada):
        # hourly_irradiance y hourly_temperature quedan en 0, como antes.
    
    # Calcular promedios por hora
    for hour in range(24):
        if hourly_counts[hour] > 0:
            hourly_ac_power[hour] /= hourly_counts[hour]
            hourly_dc_power[hour] /= hourly_counts[hour]
            hourly_irradiance[hour] /= hourly_counts[hour]
            hourly_temperature[hour] /= hourly_counts[hour]
            
            # Calcular eficiencia por hora
            if hourly_dc_power[hour] > 0:
                hourly_efficiency[hour] = (hourly_ac_power[hour] / hourly_dc_power[hour]) * 100
            
            # Convertir potencia a energía (kWh) - asumiendo mediciones cada 2 minutos
            hourly_generation[hour] = hourly_ac_power[hour] * (2/60) / 1000
    
    return {
        'hourly_efficiency': hourly_efficiency,
        'hourly_generation': hourly_generation,
        'hourly_irradiance': hourly_irradiance,
        'hourly_temperature': hourly_temperature,
        'hourly_dc_power': hourly_dc_power,
        'hourly_ac_power': hourly_ac_power,
    }


@shared_task
def calculate_inverter_data(time_range='daily', start_date_str=None, end_date_str=None, institution_id=None, device_id=None):
    """
    Calcula datos de inversores para un rango de fechas y filtros específicos.
    
    Parámetros:
        time_range: 'daily' o 'monthly'
        start_date_str: fecha de inicio en formato ISO
        end_date_str: fecha de fin en formato ISO
        institution_id: ID de la institución (opcional)
        device_id: ID del dispositivo específico (opcional)
    """
    logger.info("=== INICIANDO TAREA: calculate_inverter_data ===")
    try:
        # Procesar fechas
        if not start_date_str or not end_date_str:
            # Por defecto, último mes
            end_date = get_colombia_date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = datetime.fromisoformat(start_date_str).date()
            end_date = datetime.fromisoformat(end_date_str).date()

        logger.info(f"Calculando datos para rango: {time_range}, desde {start_date} hasta {end_date}")

        # Obtener inversores filtrados
        inverters = Device.objects.filter(category__name__iexact='inverter', is_active=True)
        
        if institution_id:
            inverters = inverters.filter(institution_id=institution_id)
            logger.info(f"Filtrado por institución ID: {institution_id}")
        
        if device_id:
            inverters = inverters.filter(scada_id=device_id)
            logger.info(f"Filtrado por dispositivo ID: {device_id}")

        logger.info(f"Procesando {inverters.count()} inversores")

        total_records_created = 0
        total_records_updated = 0

        for inverter in inverters:
            logger.info(f"Procesando inversor: {inverter.name} (ID: {inverter.id}, SCADA ID: {inverter.scada_id})")
            
            if time_range == 'daily':
                records_created, records_updated = _calculate_daily_inverter_data(inverter, start_date, end_date)
            else:  # monthly
                records_created, records_updated = _calculate_monthly_inverter_data(inverter, start_date, end_date)
            
            total_records_created += records_created
            total_records_updated += records_updated

        logger.info(f"=== RESUMEN DE PROCESAMIENTO ===")
        logger.info(f"Registros creados: {total_records_created}")
        logger.info(f"Registros actualizados: {total_records_updated}")
        logger.info("=== TAREA COMPLETADA: calculate_inverter_data ===")

        return f"Procesados {inverters.count()} inversores. Creados: {total_records_created}, Actualizados: {total_records_updated}"

    except Exception as e:
        logger.error(f"=== ERROR EN TAREA: calculate_inverter_data ===")
        logger.error(f"Error calculando datos de inversores: {e}", exc_info=True)
        raise


def _calculate_daily_inverter_data(inverter, start_date, end_date):
    """Datos diarios de un inversor (itera días delegando en el cálculo por día)."""
    return run_over_days(inverter, start_date, end_date, calculate_inverter_indicators)


def _calculate_monthly_inverter_data(inverter, start_date, end_date):
    """Datos mensuales de un inversor (itera meses delegando en el cálculo mensual)."""
    return run_over_months(inverter, start_date, end_date, calculate_inverter_indicators)


@shared_task(bind=True, retry_backoff=60, max_retries=3)
def calculate_electrical_data(self, time_range='daily', start_date_str=None, end_date_str=None, institution_id=None, device_id=None):
    """
    Calcula y actualiza indicadores eléctricos para un rango de fechas específico.
    Similar a calculate_inverter_data pero para medidores eléctricos.
    
    Args:
        time_range (str): 'daily' o 'monthly'
        start_date_str (str): Fecha de inicio en formato 'YYYY-MM-DD'
        end_date_str (str): Fecha de fin en formato 'YYYY-MM-DD'
        institution_id (int): ID de la institución (opcional)
        device_id (str): SCADA ID del dispositivo (opcional)
    """
    logger.info("=== INICIANDO TAREA: calculate_electrical_data ===")
    try:
        # Configurar fechas por defecto si no se especifican
        if not start_date_str or not end_date_str:
            # Por defecto, último mes
            end_date = get_colombia_date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = datetime.fromisoformat(start_date_str).date()
            end_date = datetime.fromisoformat(end_date_str).date()

        logger.info(f"Calculando datos eléctricos para rango: {time_range}, desde {start_date} hasta {end_date}")

        # Obtener medidores eléctricos filtrados
        electric_meters = Device.objects.filter(category__name__iexact='electricMeter', is_active=True)
        
        if institution_id:
            electric_meters = electric_meters.filter(institution_id=institution_id)
            logger.info(f"Filtrado por institución ID: {institution_id}")
        
        if device_id:
            electric_meters = electric_meters.filter(scada_id=device_id)
            logger.info(f"Filtrado por dispositivo ID: {device_id}")

        logger.info(f"Procesando {electric_meters.count()} medidores eléctricos")

        total_records_created = 0
        total_records_updated = 0

        for meter in electric_meters:
            logger.info(f"Procesando medidor: {meter.name} (ID: {meter.id}, SCADA ID: {meter.scada_id})")
            
            if time_range == 'daily':
                records_created, records_updated = _calculate_daily_electrical_data(meter, start_date, end_date)
            else:  # monthly
                records_created, records_updated = _calculate_monthly_electrical_data(meter, start_date, end_date)
            
            total_records_created += records_created
            total_records_updated += records_updated

        logger.info(f"=== RESUMEN DE PROCESAMIENTO ELÉCTRICO ===")
        logger.info(f"Registros creados: {total_records_created}")
        logger.info(f"Registros actualizados: {total_records_updated}")
        logger.info("=== TAREA COMPLETADA: calculate_electrical_data ===")

        return f"Procesados {electric_meters.count()} medidores eléctricos. Creados: {total_records_created}, Actualizados: {total_records_updated}"

    except Exception as e:
        logger.error(f"=== ERROR EN TAREA: calculate_electrical_data ===")
        logger.error(f"Error calculando datos eléctricos: {e}", exc_info=True)
        raise


def _calculate_daily_electrical_data(meter, start_date, end_date):
    """Datos diarios de un medidor eléctrico (itera días delegando en el cálculo por día)."""
    return run_over_days(meter, start_date, end_date, calculate_electric_meter_indicators)


def _calculate_monthly_electrical_data(meter, start_date, end_date):
    """Datos mensuales de un medidor eléctrico (itera meses delegando en el cálculo mensual)."""
    return run_over_months(meter, start_date, end_date, calculate_electric_meter_indicators)

@shared_task
def calculate_weather_station_indicators(time_range='daily', start_date_str=None, end_date_str=None, institution_id=None, device_id=None):
    """
    Calcula los indicadores meteorológicos para estaciones meteorológicas
    en diferentes rangos de tiempo (diario/mensual).
    
    Parámetros:
        time_range: 'daily' o 'monthly'
        start_date_str: fecha de inicio en formato ISO
        end_date_str: fecha de fin en formato ISO
        institution_id: ID de la institución (opcional)
        device_id: ID del dispositivo específico (opcional)
    """
    logger.info("=== INICIANDO TAREA: calculate_weather_station_indicators ===")
    logger.info(f"Parámetros: time_range={time_range}, start_date_str={start_date_str}, end_date_str={end_date_str}, institution_id={institution_id}, device_id={device_id}")
    
    try:
        # Paso 1: Validar y preparar fechas
        if not start_date_str or not end_date_str:
            # Por defecto, último mes
            end_date = get_colombia_date()
            if time_range == 'daily':
                start_date = end_date
            else:  # monthly
                start_date = end_date.replace(day=1)
        else:
            start_date = datetime.fromisoformat(start_date_str).date()
            end_date = datetime.fromisoformat(end_date_str).date()
        
        logger.info(f"Calculando datos para rango: {time_range}, desde {start_date} hasta {end_date}")
        
        # Paso 2: Obtener dispositivos
        # Filtrar por categoría de estación meteorológica (category_id=3 según variables.json)
        weather_stations = Device.objects.filter(category__name__iexact='weatherStation', is_active=True)
        
        if institution_id:
            weather_stations = weather_stations.filter(institution_id=institution_id)
            logger.info(f"Filtrado por institución ID: {institution_id}")
        
        if device_id:
            weather_stations = weather_stations.filter(scada_id=device_id)
            logger.info(f"Filtrado por dispositivo ID: {device_id}")
        
        logger.info(f"Procesando {weather_stations.count()} estaciones meteorológicas")
        
        total_records_created = 0
        total_records_updated = 0
        
        for station in weather_stations:
            logger.info(f"Procesando estación: {station.name} (ID: {station.id}, SCADA ID: {station.scada_id})")
            
            if time_range == 'daily':
                records_created, records_updated = _calculate_daily_weather_station_data(station, start_date, end_date)
            else:  # monthly
                records_created, records_updated = _calculate_monthly_weather_station_data(station, start_date, end_date)
            
            total_records_created += records_created
            total_records_updated += records_updated
        
        logger.info(f"=== RESUMEN DE PROCESAMIENTO ===")
        logger.info(f"Registros creados: {total_records_created}")
        logger.info(f"Registros actualizados: {total_records_updated}")
        logger.info("=== TAREA COMPLETADA: calculate_weather_station_indicators ===")
        
        return f"Procesadas {weather_stations.count()} estaciones meteorológicas. Creados: {total_records_created}, Actualizados: {total_records_updated}"
        
    except Exception as e:
        logger.error(f"=== ERROR EN TAREA: calculate_weather_station_indicators ===")
        logger.error(f"Error calculando datos de estaciones meteorológicas: {e}", exc_info=True)
        raise
        



def calculate_daily_weather_indicators(station, measurements, start_date, end_date):
    """
    Calcula indicadores meteorológicos diarios para una estación específica.
    """
    total_processed = 0
    
    # Agrupar mediciones por día (filas dict v2: measurement['date'])
    daily_measurements = {}
    for measurement in measurements:
        day = measurement['date'].date()
        if day not in daily_measurements:
            daily_measurements[day] = []
        daily_measurements[day].append(measurement)
    
    for day, day_measurements in daily_measurements.items():
        try:
            # Calcular indicadores para el día
            indicators = calculate_single_day_weather_indicators(day_measurements)
            
            # Guardar o actualizar indicadores
            weather_indicator, created = WeatherStationIndicators.objects.update_or_create(
                device=station,
                institution=station.institution,
                date=day,
                time_range='daily',
                defaults=indicators
            )
            
            # Calcular y guardar datos de gráficos
            chart_data = calculate_single_day_weather_chart_data(day_measurements)
            weather_chart, chart_created = WeatherStationChartData.objects.update_or_create(
                device=station,
                institution=station.institution,
                date=day,
                defaults=chart_data
            )
            
            total_processed += 1
            logger.info(f"Indicadores diarios calculados para {station.name} - {day}")
            
        except Exception as e:
            logger.error(f"Error calculando indicadores diarios para {station.name} - {day}: {str(e)}")
            continue
    
    return total_processed


def calculate_monthly_weather_indicators(station, measurements, start_date, end_date):
    """
    Calcula indicadores meteorológicos mensuales para una estación específica.
    """
    total_processed = 0
    
    # Agrupar mediciones por mes (filas dict v2: measurement['date'])
    monthly_measurements = {}
    for measurement in measurements:
        month_start = measurement['date'].date().replace(day=1)
        if month_start not in monthly_measurements:
            monthly_measurements[month_start] = []
        monthly_measurements[month_start].append(measurement)
    
    for month_start, month_measurements in monthly_measurements.items():
        try:
            # Calcular indicadores para el mes
            indicators = calculate_single_month_weather_indicators(month_measurements)
            
            # Guardar o actualizar indicadores
            weather_indicator, created = WeatherStationIndicators.objects.update_or_create(
                device=station,
                institution=station.institution,
                date=month_start,
                time_range='monthly',
                defaults=indicators
            )
            
            total_processed += 1
            logger.info(f"Indicadores mensuales calculados para {station.name} - {month_start}")
            
        except Exception as e:
            logger.error(f"Error calculando indicadores mensuales para {station.name} - {month_start}: {str(e)}")
            continue
    
    return total_processed


def calculate_single_day_weather_indicators(measurements):
    """
    Calcula indicadores meteorológicos para un día específico.

    Recibe filas dict (v2: `.values('date', <columnas meteorológicas>)`);
    una columna NULL equivale a la clave ausente del antiguo jsonb.
    """
    if not measurements:
        return {}

    # Extraer datos de las mediciones
    irradiance_values = []
    temperature_values = []
    humidity_values = []
    wind_speed_values = []
    wind_direction_values = []
    precipitation_values = []

    # Se descartan lecturas fuera de rango físico (sensores saturados/averiados que
    # corrompían los promedios: irradiancia negativa o >1968 W/m², viento >400 km/h).
    for data in measurements:
        # Irradiancia (W/m²): rango físico 0–1100 y solo en horas de luz (06–18 local).
        # Fuera de ese horario un valor alto es un sensor pegado que inflaba la
        # acumulación diaria (llegaba a ~24 kWh/m² vs ~8 físico). Nariño es ~1°N, con
        # día solar estable ~06–18 todo el año, así que la ventana no recorta energía real.
        irr = data['irradiance']
        if irr is not None and 0.0 <= float(irr) <= 1100.0:
            ts = data['date']
            try:
                hour = ts.astimezone(COLOMBIA_TZ).hour
            except (ValueError, AttributeError):
                hour = 12  # sin tz utilizable: no descartar por hora
            if 6 <= hour < 18:
                irradiance_values.append(float(irr))

        # Temperatura (°C): rango plausible -20–60
        temp = data['temperature']
        if temp is not None and -20.0 <= float(temp) <= 60.0:
            temperature_values.append(float(temp))

        # Humedad (%): 0–100
        hum = data['humidity']
        if hum is not None and 0.0 <= float(hum) <= 100.0:
            humidity_values.append(float(hum))

        # Velocidad del viento (km/h): 0–150
        ws = data['windSpeed']
        if ws is not None and 0.0 <= float(ws) <= 150.0:
            wind_speed_values.append(float(ws))

        # Dirección del viento (°): 0–360
        wd = data['windDirection']
        if wd is not None and 0.0 <= float(wd) <= 360.0:
            wind_direction_values.append(float(wd))

        # Precipitación (cm/día): no negativa
        prec = data['precipitation']
        if prec is not None and float(prec) >= 0.0:
            precipitation_values.append(float(prec))
    
    # Calcular indicadores
    indicators = {}
    
    # 5.1. Irradiancia Acumulada Diaria (kWh/m²)
    if irradiance_values:
        # Fórmula: Suma de irradiance (W/m²) × (2/60) horas × (1/1000) kW/W
        # Cada lectura de 2 minutos se convierte a energía: W/m² × (2/60) h = Wh/m²
        # Luego se convierte a kWh/m² dividiendo por 1000
        total_irradiance_wh_m2 = sum(irradiance_values) * (2/60)  # Wh/m²
        indicators['daily_irradiance_kwh_m2'] = total_irradiance_wh_m2 / 1000  # kWh/m²
        
        # 5.2. Horas Solares Pico (HSP)
        # 1 HSP = 1 kWh/m²
        indicators['daily_hsp_hours'] = indicators['daily_irradiance_kwh_m2']
        
        # 5.5. Generación Fotovoltaica Potencia (teórica)
        # Potencia instantánea basada en irradiancia actual
        # Asumiendo eficiencia típica del 15-20% y área de 1 m²
        efficiency = 0.17  # 17% de eficiencia típica
        # Usar la irradiancia promedio del día para calcular potencia teórica
        avg_irradiance_wm2 = sum(irradiance_values) / len(irradiance_values)
        indicators['theoretical_pv_power_w'] = avg_irradiance_wm2 * efficiency  # W instantáneos
    
    # 5.3. Viento: Velocidad Media
    if wind_speed_values:
        indicators['avg_wind_speed_kmh'] = sum(wind_speed_values) / len(wind_speed_values)
        
        # Rosa de los vientos (distribución de direcciones)
        if wind_direction_values:
            indicators['wind_direction_distribution'] = calculate_wind_direction_distribution(wind_direction_values)
            indicators['wind_speed_distribution'] = calculate_wind_speed_distribution(wind_speed_values)
    
    # 5.4. Precipitación Acumulada
    if precipitation_values:
        # Si precipitation ya está en cm/día (acumulador diario), tomar el último valor
        # Si es una tasa instantánea, se deben sumar las lecturas de 2 minutos
        # Asumiendo que cm/día significa que es un acumulador de reinicio diario
        indicators['daily_precipitation_cm'] = precipitation_values[len(precipitation_values)-1] if precipitation_values else 0.0
    
    # Datos adicionales
    if temperature_values:
        indicators['avg_temperature_c'] = sum(temperature_values) / len(temperature_values)
        indicators['max_temperature_c'] = max(temperature_values)
        indicators['min_temperature_c'] = min(temperature_values)
    
    if humidity_values:
        indicators['avg_humidity_pct'] = sum(humidity_values) / len(humidity_values)
    
    # Metadatos
    indicators['measurement_count'] = len(measurements)
    indicators['last_measurement_date'] = measurements[len(measurements)-1]['date'] if measurements else None

    return indicators


def calculate_single_month_weather_indicators(measurements):
    """
    Calcula indicadores meteorológicos para un mes específico.
    """
    if not measurements:
        return {}
    
    # Agrupar por día y calcular promedios mensuales (filas dict v2)
    daily_indicators = []
    current_day = None
    day_measurements = []

    for measurement in measurements:
        day = measurement['date'].date()
        if current_day != day:
            if day_measurements:
                daily_indicators.append(calculate_single_day_weather_indicators(day_measurements))
            current_day = day
            day_measurements = []
        day_measurements.append(measurement)
    
    # Procesar el último día
    if day_measurements:
        daily_indicators.append(calculate_single_day_weather_indicators(day_measurements))
    
    if not daily_indicators:
        return {}
    
    # Calcular promedios mensuales
    monthly_indicators = {}
    
    # Promedios de valores diarios (magnitudes intensivas: irradiancia media diaria,
    # HSP, viento, temperatura, humedad)
    for field in ['daily_irradiance_kwh_m2', 'daily_hsp_hours', 'avg_wind_speed_kmh', 'avg_temperature_c', 'avg_humidity_pct']:
        values = [ind.get(field, 0) for ind in daily_indicators if ind.get(field) is not None]
        if values:
            monthly_indicators[field] = sum(values) / len(values)

    # Precipitación mensual = ACUMULADA (suma de los diarios), no promedio
    # (indicators.md la define como acumulada). Antes se promediaba, subestimándola.
    precip_values = [ind.get('daily_precipitation_cm', 0) for ind in daily_indicators
                     if ind.get('daily_precipitation_cm') is not None]
    if precip_values:
        monthly_indicators['daily_precipitation_cm'] = sum(precip_values)
    
    # Valores máximos y mínimos
    for field in ['max_temperature_c', 'min_temperature_c']:
        values = [ind.get(field, 0) for ind in daily_indicators if ind.get(field) is not None]
        if values:
            if 'max' in field:
                monthly_indicators[field] = max(values)
            else:
                monthly_indicators[field] = min(values)
    
    # Metadatos
    monthly_indicators['measurement_count'] = sum(ind.get('measurement_count', 0) for ind in daily_indicators)
    monthly_indicators['last_measurement_date'] = measurements[len(measurements)-1]['date'] if measurements else None
    
    return monthly_indicators


def calculate_single_day_weather_chart_data(measurements):
    """
    Calcula datos de gráficos para un día específico.

    Recibe filas dict (v2: `.values('date', <columnas meteorológicas>)`).
    """
    if not measurements:
        return {}

    # Agrupar mediciones por hora
    hourly_data = {i: {'irradiance': [], 'temperature': [], 'humidity': [], 'wind_speed': [], 'wind_direction': [], 'precipitation': []} for i in range(24)}

    for data in measurements:
        hour = data['date'].hour

        if data['irradiance'] is not None:
            hourly_data[hour]['irradiance'].append(float(data['irradiance']))
        if data['temperature'] is not None:
            hourly_data[hour]['temperature'].append(float(data['temperature']))
        if data['humidity'] is not None:
            hourly_data[hour]['humidity'].append(float(data['humidity']))
        if data['windSpeed'] is not None:
            hourly_data[hour]['wind_speed'].append(float(data['windSpeed']))
        if data['windDirection'] is not None:
            hourly_data[hour]['wind_direction'].append(float(data['windDirection']))
        if data['precipitation'] is not None:
            hourly_data[hour]['precipitation'].append(float(data['precipitation']))
    
    # Calcular promedios por hora
    chart_data = {
        'hourly_irradiance': [],
        'hourly_temperature': [],
        'hourly_humidity': [],
        'hourly_wind_speed': [],
        'hourly_wind_direction': [],
        'hourly_precipitation': []
    }
    
    for hour in range(24):
        # Irradiancia promedio por hora
        if hourly_data[hour]['irradiance']:
            chart_data['hourly_irradiance'].append(sum(hourly_data[hour]['irradiance']) / len(hourly_data[hour]['irradiance']))
        else:
            chart_data['hourly_irradiance'].append(0)
        
        # Temperatura promedio por hora
        if hourly_data[hour]['temperature']:
            chart_data['hourly_temperature'].append(sum(hourly_data[hour]['temperature']) / len(hourly_data[hour]['temperature']))
        else:
            chart_data['hourly_temperature'].append(0)
        
        # Humedad promedio por hora
        if hourly_data[hour]['humidity']:
            chart_data['hourly_humidity'].append(sum(hourly_data[hour]['humidity']) / len(hourly_data[hour]['humidity']))
        else:
            chart_data['hourly_humidity'].append(0)
        
        # Velocidad del viento promedio por hora
        if hourly_data[hour]['wind_speed']:
            chart_data['hourly_wind_speed'].append(sum(hourly_data[hour]['wind_speed']) / len(hourly_data[hour]['wind_speed']))
        else:
            chart_data['hourly_wind_speed'].append(0)
        
        # Dirección del viento promedio por hora
        if hourly_data[hour]['wind_direction']:
            chart_data['hourly_wind_direction'].append(sum(hourly_data[hour]['wind_direction']) / len(hourly_data[hour]['wind_direction']))
        else:
            chart_data['hourly_wind_direction'].append(0)
        
        # Precipitación acumulada por hora (si es acumulador)
        if hourly_data[hour]['precipitation']:
            chart_data['hourly_precipitation'].append(hourly_data[hour]['precipitation'][len(hourly_data[hour]['precipitation'])-1])  # Último valor
        else:
            chart_data['hourly_precipitation'].append(0)
    
    # Calcular valores diarios
    chart_data['daily_irradiance_kwh_m2'] = sum(chart_data['hourly_irradiance']) * (2/60) / 1000  # kWh/m²
    chart_data['avg_daily_temperature_c'] = sum(chart_data['hourly_temperature']) / 24
    chart_data['avg_daily_humidity_pct'] = sum(chart_data['hourly_humidity']) / 24
    chart_data['avg_daily_wind_speed_kmh'] = sum(chart_data['hourly_wind_speed']) / 24
    chart_data['daily_precipitation_cm'] = chart_data['hourly_precipitation'][len(chart_data['hourly_precipitation'])-1] if chart_data['hourly_precipitation'] else 0
    
    return chart_data


def calculate_wind_direction_distribution(wind_directions):
    """
    Calcula la distribución de direcciones del viento para la rosa de los vientos.
    """
    # Dividir en 8 direcciones principales (N, NE, E, SE, S, SW, W, NW)
    directions = ['N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW']
    direction_bins = {direction: 0 for direction in directions}
    
    for direction in wind_directions:
        # Convertir grados a dirección cardinal
        if 337.5 <= direction <= 360 or 0 <= direction < 22.5:
            direction_bins['N'] += 1
        elif 22.5 <= direction < 67.5:
            direction_bins['NE'] += 1
        elif 67.5 <= direction < 112.5:
            direction_bins['E'] += 1
        elif 112.5 <= direction < 157.5:
            direction_bins['SE'] += 1
        elif 157.5 <= direction < 202.5:
            direction_bins['S'] += 1
        elif 202.5 <= direction < 247.5:
            direction_bins['SW'] += 1
        elif 247.5 <= direction < 292.5:
            direction_bins['W'] += 1
        elif 292.5 <= direction < 337.5:
            direction_bins['NW'] += 1
    
    return direction_bins


def calculate_wind_speed_distribution(wind_speeds):
    """
    Calcula la distribución de velocidades del viento.
    """
    # Definir rangos de velocidad
    speed_ranges = {
        '0-5': 0,    # Calma
        '5-10': 0,   # Ligera
        '10-20': 0,  # Moderada
        '20-30': 0,  # Fuerte
        '30+': 0     # Muy fuerte
    }
    
    for speed in wind_speeds:
        if speed < 5:
            speed_ranges['0-5'] += 1
        elif speed < 10:
            speed_ranges['5-10'] += 1
        elif speed < 20:
            speed_ranges['10-20'] += 1
        elif speed < 30:
            speed_ranges['20-30'] += 1
        else:
            speed_ranges['30+'] += 1
    
    return speed_ranges


def _calculate_daily_weather_station_data(station, start_date, end_date):
    """
    Calcula datos diarios para una estación meteorológica específica
    """
    records_created = 0
    records_updated = 0
    
    current_date = start_date
    while current_date <= end_date:
        logger.info(f"  Procesando fecha: {current_date}")
        
        try:
            # Obtener mediciones para el día específico (v2: columnas tipadas,
            # rango aware Bogotá equivalente al antiguo date__date)
            day_a, day_b = colombia_day_range(current_date, current_date)
            measurements = WeatherStationMeasurement.objects.filter(
                device=station,
                date__gte=day_a, date__lt=day_b
            ).order_by('date')

            if measurements.exists():
                # Filas dict con las columnas usadas por los cálculos
                measurements_list = list(measurements.values(
                    'date', 'irradiance', 'temperature', 'humidity',
                    'windSpeed', 'windDirection', 'precipitation',
                ))
                
                # Calcular indicadores para el día
                indicators = calculate_single_day_weather_indicators(measurements_list)
                
                # Guardar o actualizar indicadores
                weather_indicator, created = WeatherStationIndicators.objects.update_or_create(
                    device=station,
                    institution=station.institution,
                    date=current_date,
                    time_range='daily',
                    defaults=indicators
                )
                
                if created:
                    records_created += 1
                else:
                    records_updated += 1
                
                # Calcular y guardar datos de gráficos
                chart_data = calculate_single_day_weather_chart_data(measurements_list)
                weather_chart, chart_created = WeatherStationChartData.objects.update_or_create(
                    device=station,
                    institution=station.institution,
                    date=current_date,
                    defaults=chart_data
                )
                
                logger.info(f"  Indicadores diarios calculados para {station.name} - {current_date}")
            else:
                logger.warning(f"  No hay mediciones para {station.name} en {current_date}")
                
        except Exception as e:
            logger.error(f"  Error calculando indicadores diarios para {station.name} - {current_date}: {str(e)}")
        
        current_date += timedelta(days=1)

    return records_created, records_updated


def _calculate_monthly_weather_station_data(station, start_date, end_date):
    """
    Calcula datos mensuales para una estación meteorológica específica
    """
    records_created = 0
    records_updated = 0
    
    # Agrupar por mes
    current_date = start_date.replace(day=1)  # Primer día del mes
    while current_date <= end_date:
        month_end = (current_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        month_end = min(month_end, end_date)
        
        logger.info(f"  Procesando mes: {current_date.strftime('%Y-%m')}")
        
        try:
            # Obtener mediciones para el mes (v2: columnas tipadas, rango aware
            # Bogotá equivalente al antiguo date__date__range)
            month_a, month_b = colombia_day_range(current_date, month_end)
            measurements = WeatherStationMeasurement.objects.filter(
                device=station,
                date__gte=month_a, date__lt=month_b
            ).order_by('date')

            if measurements.exists():
                # Filas dict con las columnas usadas por los cálculos
                measurements_list = list(measurements.values(
                    'date', 'irradiance', 'temperature', 'humidity',
                    'windSpeed', 'windDirection', 'precipitation',
                ))
                
                # Calcular indicadores para el mes
                indicators = calculate_single_month_weather_indicators(measurements_list)
                
                # Guardar o actualizar indicadores
                weather_indicator, created = WeatherStationIndicators.objects.update_or_create(
                    device=station,
                    institution=station.institution,
                    date=current_date,
                    time_range='monthly',
                    defaults=indicators
                )
                
                if created:
                    records_created += 1
                else:
                    records_updated += 1
                
                logger.info(f"  Indicadores mensuales calculados para {station.name} - {current_date.strftime('%Y-%m')}")
            else:
                logger.warning(f"  No hay mediciones para {station.name} en {current_date.strftime('%Y-%m')}")
                
        except Exception as e:
            logger.error(f"  Error calculando indicadores mensuales para {station.name} - {current_date.strftime('%Y-%m')}: {str(e)}")
        
        # Avanzar al siguiente mes
        current_date = (current_date.replace(day=1) + timedelta(days=32)).replace(day=1)

    return records_created, records_updated

# =========================
# TAREAS PARA GENERACIÓN DE REPORTE
# =========================

@shared_task(bind=True, retry_backoff=60, max_retries=3,
             soft_time_limit=1800, time_limit=2000)
def generate_report(self, institution_id, category, devices, report_type, time_range, start_date, end_date, format, user_id):
    """
    Genera un reporte en el formato especificado
    """
    logger.info(f"=== INICIANDO GENERACIÓN DE REPORTE ===")
    logger.info(f"Tarea ID: {self.request.id}")
    logger.info(f"Parámetros: {institution_id}, {category}, {report_type}, {time_range}, {start_date} - {end_date}, {format}")
    
    try:
        # Crear registro del reporte
        from .models import GeneratedReport
        from scada_proxy.models import Institution
        
        # Obtener nombre de la institución
        try:
            institution = Institution.objects.get(id=institution_id)
            institution_name = institution.name
        except Institution.DoesNotExist:
            raise ValueError(f"Institución con ID {institution_id} no encontrada")
        
        # Crear o actualizar registro del reporte
        report_record, created = GeneratedReport.objects.get_or_create(
            task_id=self.request.id,
            defaults={
                'user_id': user_id,
                'report_type': report_type,
                'category': category,
                'institution_id': institution_id,
                'institution_name': institution_name,
                'devices': devices,
                'time_range': time_range,
                'start_date': start_date,
                'end_date': end_date,
                'format': format,
                'status': 'processing'
            }
        )
        
        if not created:
            report_record.status = 'processing'
            report_record.save()
        
        logger.info(f"Registro de reporte {'creado' if created else 'actualizado'} con ID: {report_record.id}")
        
        # Actualizar progreso de la tarea
        self.update_state(
            state='PROGRESS',
            meta={'current': 0, 'total': 100, 'status': 'Iniciando generación de reporte'}
        )
        
        logger.info(f"Estado de tarea actualizado a PROGRESS")
        
        # Generar reporte según la categoría y tipo
        logger.info(f"Generando reporte para categoría: {category}")
        
        if category == 'electricMeter':
            report_data = generate_electric_meter_report(
                institution_id, devices, report_type, time_range, start_date, end_date
            )
        elif category == 'inverter':
            report_data = generate_inverter_report(
                institution_id, devices, report_type, time_range, start_date, end_date
            )
        elif category == 'weatherStation':
            report_data = generate_weather_station_report(
                institution_id, devices, report_type, time_range, start_date, end_date
            )
        else:
            raise ValueError(f"Categoría no válida: {category}")
        
        logger.info(f"Reporte generado exitosamente para categoría {category}")
        
        # Actualizar progreso
        self.update_state(
            state='PROGRESS',
            meta={'current': 50, 'total': 100, 'status': 'Generando archivo'}
        )
        
        logger.info(f"Estado de tarea actualizado a 50%")
        
        # Generar archivo en el formato especificado
        logger.info(f"Generando archivo en formato: {format}")
        
        file_path, file_size, record_count = generate_report_file(
            report_data, report_type, format, self.request.id
        )
        
        logger.info(f"Archivo generado: {file_path}, tamaño: {file_size}, registros: {record_count}")
        
        # Actualizar progreso
        self.update_state(
            state='PROGRESS',
            meta={'current': 90, 'total': 100, 'status': 'Finalizando'}
        )
        
        logger.info(f"Estado de tarea actualizado a 90%")
        
        # Actualizar registro del reporte
        report_record.status = 'completed'
        report_record.file_path = file_path
        report_record.file_size = file_size
        report_record.record_count = record_count
        report_record.completed_at = django_timezone.now()
        report_record.save()
        
        logger.info(f"Registro de reporte actualizado a completado")
        logger.info(f"Reporte generado exitosamente: {file_path}")
        
        return {
            'status': 'completed',
            'file_path': file_path,
            'file_size': file_size,
            'record_count': record_count
        }
        
    except Exception as e:
        logger.error(f"Error generando reporte: {str(e)}")
        
        # Actualizar registro del reporte con error
        try:
            report_record = GeneratedReport.objects.get(task_id=self.request.id)
            report_record.status = 'failed'
            report_record.error_message = str(e)
            report_record.save()
        except GeneratedReport.DoesNotExist:
            pass
        
        # Reintentar si es posible
        if self.request.retries < self.max_retries:
            raise self.retry(exc=e, countdown=60)
        else:
            raise e


def generate_electric_meter_report(institution_id, devices, report_type, time_range, start_date, end_date):
    """
    Genera datos para reportes de medidores eléctricos
    """
    from .models import ElectricMeterIndicators
    
    logger.info(f"Generando reporte de medidores eléctricos: {report_type}")
    logger.info(f"Parámetros: institution_id={institution_id}, devices={devices}, time_range={time_range}, start_date={start_date}, end_date={end_date}")
    
    # Convertir fechas
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Construir filtros base
    filters = Q(
        institution_id=institution_id,
        time_range=time_range,
        date__range=(start_date, end_date)
    )
    
    # Validar y filtrar dispositivos
    if devices and isinstance(devices, list) and len(devices) > 0:
        # Asegurar que todos los IDs sean números
        try:
            device_ids = [int(device_id) for device_id in devices if device_id is not None]
            if device_ids:
                filters &= Q(device_id__in=device_ids)
                logger.info(f"Filtro de dispositivos aplicado: {device_ids}")
            else:
                logger.info("No se aplicó filtro de dispositivos - lista vacía después de conversión")
        except (ValueError, TypeError) as e:
            logger.warning(f"Error convirtiendo IDs de dispositivos: {e}. Dispositivos: {devices}")
    else:
        logger.info("No se aplicó filtro de dispositivos - parámetro vacío o inválido")
    
    # Obtener datos según el tipo de reporte
    if report_type == 'Resumen de Consumo':
        # Repuntado a ElectricMeterIndicators (fuente única, energía saneada anti
        # roll-over) en vez de ElectricMeterEnergyConsumption, que estaba vacía.
        data = ElectricMeterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'imported_energy_kwh', 'exported_energy_kwh', 'net_energy_consumption_kwh']
        
    elif report_type == 'Análisis de Demanda':
        data = ElectricMeterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'peak_demand_kw', 'avg_demand_kw', 'load_factor_pct']
        
    elif report_type == 'Calidad de Potencia':
        data = ElectricMeterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'max_voltage_thd_pct', 'max_current_thd_pct', 'avg_power_factor']
        
    elif report_type == 'Balance Energético':
        # Repuntado a ElectricMeterIndicators (energía saneada) en vez de la tabla vacía.
        data = ElectricMeterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'imported_energy_kwh', 'exported_energy_kwh', 'net_energy_consumption_kwh']
        
    elif report_type == 'Reporte Integral':
        data = ElectricMeterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'imported_energy_kwh', 'exported_energy_kwh', 'peak_demand_kw', 'avg_power_factor', 'max_voltage_thd_pct']
        
    else:
        raise ValueError(f"Tipo de reporte no válido: {report_type}")
    
    # Preparar datos para el reporte
    report_data = []
    for item in data:
        row = {}
        for col in columns:
            if '__' in col:
                # Campo relacionado
                parts = col.split('__')
                value = item
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        break
            else:
                # Campo directo
                value = getattr(item, col, None)
            
            # Formatear valor
            if isinstance(value, (int, float)):
                if 'pct' in col or 'factor' in col:
                    row[col.split('__')[-1]] = f"{value:.2f}%"
                elif 'kw' in col or 'kwh' in col:
                    row[col.split('__')[-1]] = f"{value:.2f}"
                else:
                    row[col.split('__')[-1]] = value
            else:
                row[col.split('__')[-1]] = value
        
        report_data.append(row)
    
    return report_data


def generate_inverter_report(institution_id, devices, report_type, time_range, start_date, end_date):
    """
    Genera datos para reportes de inversores
    """
    from .models import InverterIndicators
    
    logger.info(f"Generando reporte de inversores: {report_type}")
    logger.info(f"Parámetros: institution_id={institution_id}, devices={devices}, time_range={time_range}, start_date={start_date}, end_date={end_date}")
    
    # Convertir fechas
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Construir filtros base
    filters = Q(
        institution_id=institution_id,
        time_range=time_range,
        date__range=(start_date, end_date)
    )
    
    # Validar y filtrar dispositivos
    if devices and isinstance(devices, list) and len(devices) > 0:
        # Asegurar que todos los IDs sean números
        try:
            device_ids = [int(device_id) for device_id in devices if device_id is not None]
            if device_ids:
                filters &= Q(device_id__in=device_ids)
                logger.info(f"Filtro de dispositivos aplicado: {device_ids}")
            else:
                logger.info("No se aplicó filtro de dispositivos - lista vacía después de conversión")
        except (ValueError, TypeError) as e:
            logger.warning(f"Error convirtiendo IDs de dispositivos: {e}. Dispositivos: {devices}")
    else:
        logger.info("No se aplicó filtro de dispositivos - parámetro vacío o inválido")
    
    # Obtener datos según el tipo de reporte
    if report_type == 'Resumen de Generación':
        data = InverterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'total_generated_energy_kwh', 'dc_ac_efficiency_pct', 'energy_ac_daily_kwh']
        
    elif report_type == 'Análisis de Rendimiento':
        data = InverterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'performance_ratio_pct', 'avg_irradiance_wm2', 'avg_temperature_c']
        
    elif report_type == 'Métricas Operativas':
        data = InverterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'avg_power_factor_pct', 'avg_frequency_hz', 'frequency_stability_pct']
        
    elif report_type == 'Reporte de Anomalías':
        data = InverterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'anomaly_score', 'max_power_w', 'min_power_w']
        
    elif report_type == 'Reporte Integral':
        data = InverterIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'total_generated_energy_kwh', 'dc_ac_efficiency_pct', 'performance_ratio_pct', 'avg_power_factor_pct']
        
    else:
        raise ValueError(f"Tipo de reporte no válido: {report_type}")
    
    # Preparar datos para el reporte
    report_data = []
    for item in data:
        row = {}
        for col in columns:
            if '__' in col:
                # Campo relacionado
                parts = col.split('__')
                value = item
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        break
            else:
                # Campo directo
                value = getattr(item, col, None)
            
            # Formatear valor
            if isinstance(value, (int, float)):
                if 'pct' in col:
                    row[col.split('__')[-1]] = f"{value:.2f}%"
                elif 'kwh' in col or 'kw' in col or 'wm2' in col:
                    row[col.split('__')[-1]] = f"{value:.2f}"
                else:
                    row[col.split('__')[-1]] = value
            else:
                row[col.split('__')[-1]] = value
        
        report_data.append(row)
    
    return report_data


def generate_weather_station_report(institution_id, devices, report_type, time_range, start_date, end_date):
    """
    Genera datos para reportes de estaciones meteorológicas
    """
    from .models import WeatherStationIndicators
    
    logger.info(f"Generando reporte de estaciones meteorológicas: {report_type}")
    logger.info(f"Parámetros: institution_id={institution_id}, devices={devices}, time_range={time_range}, start_date={start_date}, end_date={end_date}")
    
    # Convertir fechas
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Construir filtros base
    filters = Q(
        institution_id=institution_id,
        time_range=time_range,
        date__range=(start_date, end_date)
    )
    
    # Validar y filtrar dispositivos
    if devices and isinstance(devices, list) and len(devices) > 0:
        # Asegurar que todos los IDs sean números
        try:
            device_ids = [int(device_id) for device_id in devices if device_id is not None]
            if device_ids:
                filters &= Q(device_id__in=device_ids)
                logger.info(f"Filtro de dispositivos aplicado: {device_ids}")
            else:
                logger.info("No se aplicó filtro de dispositivos - lista vacía después de conversión")
        except (ValueError, TypeError) as e:
            logger.warning(f"Error convirtiendo IDs de dispositivos: {e}. Dispositivos: {devices}")
    else:
        logger.info("No se aplicó filtro de dispositivos - parámetro vacío o inválido")
    
    # Obtener datos según el tipo de reporte
    if report_type == 'Resumen Climático':
        data = WeatherStationIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'avg_temperature_c', 'avg_humidity_pct', 'daily_precipitation_cm']
        
    elif report_type == 'Análisis Solar':
        data = WeatherStationIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'daily_irradiance_kwh_m2', 'daily_hsp_hours', 'theoretical_pv_power_w']
        
    elif report_type == 'Análisis de Viento':
        data = WeatherStationIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'avg_wind_speed_kmh', 'wind_direction_distribution', 'wind_speed_distribution']
        
    elif report_type == 'Impacto Ambiental':
        data = WeatherStationIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'avg_temperature_c', 'daily_irradiance_kwh_m2', 'theoretical_pv_power_w']
        
    elif report_type == 'Reporte Integral':
        data = WeatherStationIndicators.objects.filter(filters).select_related('device')
        columns = ['device__name', 'date', 'daily_irradiance_kwh_m2', 'avg_temperature_c', 'avg_wind_speed_kmh', 'daily_precipitation_cm']
        
    else:
        raise ValueError(f"Tipo de reporte no válido: {report_type}")
    
    # Preparar datos para el reporte
    report_data = []
    for item in data:
        row = {}
        for col in columns:
            if '__' in col:
                # Campo relacionado
                parts = col.split('__')
                value = item
                for part in parts:
                    value = getattr(value, part, None)
                    if value is None:
                        break
            else:
                # Campo directo
                value = getattr(item, col, None)
            
            # Formatear valor
            if isinstance(value, (int, float)):
                if 'pct' in col:
                    row[col.split('__')[-1]] = f"{value:.2f}%"
                elif 'kwh' in col or 'kmh' in col or 'cm' in col:
                    row[col.split('__')[-1]] = f"{value:.2f}"
                else:
                    row[col.split('__')[-1]] = value
            else:
                row[col.split('__')[-1]] = value
        
        report_data.append(row)
    
    return report_data


def generate_report_file(report_data, report_type, format, task_id):
    """
    Genera el archivo del reporte en el formato especificado
    """
    from django.conf import settings
    
    # Crear directorio para reportes si no existe
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # Generar nombre del archivo
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"reporte_{report_type.replace(' ', '_')}_{timestamp}_{task_id[:8]}"
    
    if format == 'CSV':
        file_path = os.path.join(reports_dir, f"{filename}.csv")
        file_size, record_count = generate_csv_file(report_data, file_path)
        
    elif format == 'Excel':
        file_path = os.path.join(reports_dir, f"{filename}.xlsx")
        file_size, record_count = generate_excel_file(report_data, file_path)
        
    elif format == 'PDF':
        file_path = os.path.join(reports_dir, f"{filename}.pdf")
        # Determinar si es un reporte ejecutivo basado en el tipo
        if 'executive' in report_type.lower() or 'ejecutivo' in report_type.lower():
            # Extraer información adicional si está disponible
            institution_name = ""
            period_info = ""
            if hasattr(report_data, 'get') and isinstance(report_data, dict):
                institution_name = report_data.get('institution_name', '')
                period_info = report_data.get('period_info', '')
            
            file_size, record_count = generate_executive_pdf_file(
                report_data, file_path, report_type, institution_name, period_info
            )
        else:
            file_size, record_count = generate_pdf_file(report_data, file_path, report_type)
        
    else:
        raise ValueError(f"Formato no soportado: {format}")
    
    return file_path, file_size, record_count


def generate_csv_file(report_data, file_path):
    """
    Genera archivo CSV
    """
    
    if not report_data:
        # Crear archivo vacío
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['No hay datos disponibles para el período seleccionado'])
        return "0 KB", 0
    
    # Obtener columnas del primer registro
    columns = list(report_data[0].keys())
    
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=columns)
        writer.writeheader()
        writer.writerows(report_data)
    
    # Calcular tamaño del archivo
    file_size = os.path.getsize(file_path)
    file_size_str = format_file_size(file_size)
    
    return file_size_str, len(report_data)


def generate_excel_file(report_data, file_path):
    """
    Genera archivo Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        # Fallback a CSV si no hay openpyxl
        logger.warning("openpyxl no disponible, generando CSV en su lugar")
        return generate_csv_file(report_data, file_path.replace('.xlsx', '.csv'))
    
    if not report_data:
        # Crear archivo vacío
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'No hay datos disponibles para el período seleccionado'
        wb.save(file_path)
        return "0 KB", 0
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Obtener columnas del primer registro
    columns = list(report_data[0].keys())
    
    # Escribir encabezados
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Escribir datos
    for row, data_row in enumerate(report_data, 2):
        for col, header in enumerate(columns, 1):
            ws.cell(row=row, column=col, value=data_row.get(header, ''))
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    
    # Calcular tamaño del archivo
    file_size = os.path.getsize(file_path)
    file_size_str = format_file_size(file_size)
    
    return file_size_str, len(report_data)


def generate_pdf_file(report_data, file_path, report_type):
    """
    Genera archivo PDF con formato mejorado y más detalle
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        from reportlab.pdfgen import canvas
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from datetime import datetime
        import io
        import numpy as np
    except ImportError as e:
        # Fallback a CSV si no hay reportlab o matplotlib
        logger.warning(f"Librerías no disponibles: {e}, generando CSV en su lugar")
        return generate_csv_file(report_data, file_path.replace('.pdf', '.csv'))
    
    # Crear documento PDF
    doc = SimpleDocTemplate(file_path, pagesize=A4, 
                          rightMargin=2*cm, leftMargin=2*cm, 
                          topMargin=2*cm, bottomMargin=2*cm)
    story = []
    
    # Estilos mejorados
    styles = getSampleStyleSheet()
    
    # Estilo para título principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1f2937'),
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#374151'),
        fontName='Helvetica-Bold',
        spaceBefore=25
    )
    
    # Estilo para texto normal
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12,
        alignment=TA_JUSTIFY,
        textColor=colors.HexColor('#4b5563'),
        fontName='Helvetica'
    )
    
    # Estilo para información destacada
    highlight_style = ParagraphStyle(
        'Highlight',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=15,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#059669'),
        fontName='Helvetica-Bold',
        backColor=colors.HexColor('#ecfdf5'),
        borderPadding=10,
        borderWidth=1,
        borderColor=colors.HexColor('#10b981')
    )
    
    # Encabezado del reporte
    story.append(Paragraph(f"REPORTE DETALLADO: {report_type.upper()}", title_style))
    story.append(Spacer(1, 20))
    
    # Información del reporte
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    story.append(Paragraph(f"<b>Fecha de generación:</b> {current_time}", normal_style))
    story.append(Paragraph(f"<b>Total de registros:</b> {len(report_data)}", normal_style))
    story.append(Spacer(1, 20))
    
    if not report_data:
        story.append(Paragraph("No hay datos disponibles para el período seleccionado", highlight_style))
        doc.build(story)
        
        # Calcular tamaño del archivo
        file_size = os.path.getsize(file_path)
        file_size_str = format_file_size(file_size)
        
        return file_size_str, 0
    
    # Resumen ejecutivo
    story.append(Paragraph("RESUMEN EJECUTIVO", subtitle_style))
    story.append(Paragraph("Este reporte presenta un análisis detallado de los datos recopilados durante el período especificado, incluyendo métricas clave, tendencias y análisis estadísticos relevantes.", normal_style))
    story.append(Spacer(1, 20))
    
    # Análisis estadístico básico
    if len(report_data) > 0:
        story.append(Paragraph("ANÁLISIS ESTADÍSTICO", subtitle_style))
        
        # Obtener columnas numéricas para análisis
        numeric_columns = []
        for col in report_data[0].keys():
            try:
                # Intentar convertir a número
                sample_value = report_data[0][col]
                if isinstance(sample_value, (int, float)) or (isinstance(sample_value, str) and sample_value.replace('.', '').replace('-', '').isdigit()):
                    numeric_columns.append(col)
            except:
                continue
        
        if numeric_columns:
            # Crear tabla de estadísticas
            stats_data = [['Métrica', 'Valor']]
            
            for col in numeric_columns[:5]:  # Limitar a 5 columnas para no sobrecargar
                try:
                    values = []
                    for row in report_data:
                        val = row.get(col, 0)
                        if isinstance(val, str):
                            val = float(val) if val.replace('.', '').replace('-', '').isdigit() else 0
                        if isinstance(val, (int, float)):
                            values.append(val)
                    
                    if values:
                        stats_data.append([f'Promedio {col}', f'{np.mean(values):.2f}'])
                        stats_data.append([f'Máximo {col}', f'{np.max(values):.2f}'])
                        stats_data.append([f'Mínimo {col}', f'{np.min(values):.2f}'])
                except:
                    continue
            
            if len(stats_data) > 1:
                stats_table = Table(stats_data, colWidths=[4*cm, 3*cm])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f9fafb'), colors.white]),
                ]))
                story.append(stats_table)
                story.append(Spacer(1, 20))
    
    # Generar gráficos si hay datos numéricos
    if len(report_data) > 1 and numeric_columns:
        story.append(Paragraph("GRÁFICOS Y VISUALIZACIONES", subtitle_style))
        
        # Gráfico de línea temporal para la primera columna numérica
        try:
            time_col = None
            for col in report_data[0].keys():
                if 'fecha' in col.lower() or 'date' in col.lower() or 'time' in col.lower():
                    time_col = col
                    break
            
            if time_col:
                # Crear gráfico de línea
                fig, ax = plt.subplots(figsize=(10, 6))
                
                # Preparar datos para el gráfico
                x_values = []
                y_values = []
                
                for row in report_data[:50]:  # Limitar a 50 puntos para el gráfico
                    try:
                        time_val = row.get(time_col, '')
                        if isinstance(time_val, str):
                            # Intentar parsear fecha
                            try:
                                parsed_time = datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S')
                                x_values.append(parsed_time)
                            except:
                                x_values.append(len(x_values))
                        else:
                            x_values.append(len(x_values))
                        
                        # Valor numérico para Y
                        for num_col in numeric_columns[:1]:  # Solo primera columna numérica
                            val = row.get(num_col, 0)
                            if isinstance(val, str):
                                val = float(val) if val.replace('.', '').replace('-', '').isdigit() else 0
                            y_values.append(val if isinstance(val, (int, float)) else 0)
                            break
                    except:
                        continue
                
                if len(x_values) > 1 and len(y_values) > 1:
                    ax.plot(x_values, y_values, marker='o', linewidth=2, markersize=4)
                    ax.set_title(f'Tendencia de {numeric_columns[0]}', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Tiempo', fontsize=12)
                    ax.set_ylabel(numeric_columns[0], fontsize=12)
                    ax.grid(True, alpha=0.3)
                    
                    # Formatear eje X si son fechas
                    if isinstance(x_values[0], datetime):
                        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
                        ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(x_values)//10)))
                        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45)
                    
                    plt.tight_layout()
                    
                    # Guardar gráfico en memoria
                    img_buffer = io.BytesIO()
                    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
                    img_buffer.seek(0)
                    
                    # Agregar imagen al PDF
                    story.append(Image(img_buffer, width=12*cm, height=7*cm))
                    story.append(Spacer(1, 20))
                    
                    plt.close()
        except Exception as e:
            logger.warning(f"Error generando gráfico: {e}")
    
    # Datos detallados
    story.append(Paragraph("DATOS DETALLADOS", subtitle_style))
    story.append(Paragraph("A continuación se presentan los datos completos del reporte en formato tabular:", normal_style))
    story.append(Spacer(1, 15))
    
    # Obtener columnas del primer registro
    columns = list(report_data[0].keys())
    
    # Limitar columnas si son demasiadas para el PDF
    if len(columns) > 8:
        story.append(Paragraph(f"<i>Nota: Se muestran las primeras 8 columnas de {len(columns)} totales para optimizar la visualización.</i>", normal_style))
        story.append(Spacer(1, 10))
        columns = columns[:8]
    
    # Crear tabla de datos con formato mejorado
    table_data = [columns]  # Encabezados
    
    # Agregar datos (limitar filas para no sobrecargar el PDF)
    max_rows = min(50, len(report_data))  # Máximo 50 filas
    for i, data_row in enumerate(report_data[:max_rows]):
        row = []
        for col in columns:
            value = data_row.get(col, '')
            # Truncar valores muy largos
            if isinstance(value, str) and len(value) > 30:
                value = value[:27] + '...'
            row.append(str(value))
        table_data.append(row)
    
    if len(report_data) > max_rows:
        story.append(Paragraph(f"<i>Nota: Se muestran las primeras {max_rows} filas de {len(report_data)} totales.</i>", normal_style))
        story.append(Spacer(1, 10))
    
    # Crear tabla con estilo mejorado
    table = Table(table_data, colWidths=[2.5*cm] * len(columns))
    table.setStyle(TableStyle([
        # Encabezados
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f9fafb'), colors.white]),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Pie de página con información adicional
    story.append(Paragraph("INFORMACIÓN ADICIONAL", subtitle_style))
    story.append(Paragraph("• Este reporte fue generado automáticamente por el sistema MTE - SIVE", normal_style))
    story.append(Paragraph("• Los datos están sujetos a validación y pueden ser actualizados", normal_style))
    story.append(Paragraph("• Para consultas adicionales, contacte al administrador del sistema", normal_style))
    
    # Construir el PDF
    doc.build(story)
    
    # Calcular tamaño del archivo
    file_size = os.path.getsize(file_path)
    file_size_str = format_file_size(file_size)
    
    return file_size_str, len(report_data)


def generate_executive_pdf_file(report_data, file_path, report_type, institution_name="", period_info=""):
    """
    Genera un PDF ejecutivo con formato profesional y resumido
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch, cm
        from reportlab.pdfgen import canvas
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from datetime import datetime
        import io
        import numpy as np
    except ImportError as e:
        logger.warning(f"Librerías no disponibles: {e}, usando función básica")
        return generate_pdf_file(report_data, file_path, report_type)
    
    # Crear documento PDF
    doc = SimpleDocTemplate(file_path, pagesize=A4, 
                          rightMargin=1.5*cm, leftMargin=1.5*cm, 
                          topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    
    # Estilos ejecutivos
    styles = getSampleStyleSheet()
    
    # Estilo para título principal ejecutivo
    executive_title_style = ParagraphStyle(
        'ExecutiveTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=25,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1e40af'),
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulos ejecutivos
    executive_subtitle_style = ParagraphStyle(
        'ExecutiveSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=15,
        alignment=TA_LEFT,
        textColor=colors.HexColor('#1e293b'),
        fontName='Helvetica-Bold',
        spaceBefore=20
    )
    
    # Estilo para texto ejecutivo
    executive_text_style = ParagraphStyle(
        'ExecutiveText',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=10,
        alignment=TA_JUSTIFY,
        textColor=colors.HexColor('#334155'),
        fontName='Helvetica'
    )
    
    # Estilo para métricas destacadas
    metric_style = ParagraphStyle(
        'Metric',
        parent=styles['Normal'],
        fontSize=13,
        spaceAfter=8,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#059669'),
        fontName='Helvetica-Bold',
        backColor=colors.HexColor('#f0fdf4'),
        borderPadding=8,
        borderWidth=1,
        borderColor=colors.HexColor('#22c55e')
    )
    
    # Encabezado ejecutivo
    story.append(Paragraph("MTE - SIVE", executive_title_style))
    story.append(Paragraph("SISTEMA DE MONITOREO INTEGRAL", executive_title_style))
    story.append(Spacer(1, 15))
    
    # Información del reporte ejecutivo
    story.append(Paragraph(f"<b>REPORTE EJECUTIVO:</b> {report_type.upper()}", executive_subtitle_style))
    if institution_name:
        story.append(Paragraph(f"<b>Institución:</b> {institution_name}", executive_text_style))
    if period_info:
        story.append(Paragraph(f"<b>Período:</b> {period_info}", executive_text_style))
    
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    story.append(Paragraph(f"<b>Generado:</b> {current_time}", executive_text_style))
    story.append(Paragraph(f"<b>Total de registros:</b> {len(report_data)}", executive_text_style))
    story.append(Spacer(1, 20))
    
    if not report_data:
        story.append(Paragraph("No hay datos disponibles para el período seleccionado", metric_style))
        doc.build(story)
        file_size = os.path.getsize(file_path)
        file_size_str = format_file_size(file_size)
        return file_size_str, 0
    
    # Resumen ejecutivo
    story.append(Paragraph("RESUMEN EJECUTIVO", executive_subtitle_style))
    story.append(Paragraph("Este reporte presenta un análisis ejecutivo de los datos recopilados, destacando las métricas clave y tendencias más relevantes para la toma de decisiones estratégicas.", executive_text_style))
    story.append(Spacer(1, 15))
    
    # KPIs principales
    if len(report_data) > 0:
        story.append(Paragraph("INDICADORES CLAVE (KPIs)", executive_subtitle_style))
        
        # Obtener columnas numéricas para análisis
        numeric_columns = []
        for col in report_data[0].keys():
            try:
                sample_value = report_data[0][col]
                if isinstance(sample_value, (int, float)) or (isinstance(sample_value, str) and sample_value.replace('.', '').replace('-', '').isdigit()):
                    numeric_columns.append(col)
            except:
                continue
        
        if numeric_columns:
            # Crear tabla de KPIs ejecutivos
            kpi_data = [['Indicador', 'Valor', 'Tendencia']]
            
            for col in numeric_columns[:3]:  # Solo 3 KPIs principales
                try:
                    values = []
                    for row in report_data:
                        val = row.get(col, 0)
                        if isinstance(val, str):
                            val = float(val) if val.replace('.', '').replace('-', '').isdigit() else 0
                        if isinstance(val, (int, float)):
                            values.append(val)
                    
                    if values:
                        avg_val = np.mean(values)
                        max_val = np.max(values)
                        min_val = np.min(values)
                        
                        # Determinar tendencia
                        if len(values) > 1:
                            trend = "↗️" if values[-1] > values[0] else "↘️" if values[-1] < values[0] else "➡️"
                        else:
                            trend = "➡️"
                        
                        kpi_data.append([f'{col}', f'{avg_val:.2f}', trend])
                except:
                    continue
            
            if len(kpi_data) > 1:
                kpi_table = Table(kpi_data, colWidths=[5*cm, 3*cm, 2*cm])
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
                    ('FONTSIZE', (0, 1), (-1, -1), 11),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
                ]))
                story.append(kpi_table)
                story.append(Spacer(1, 20))
    
    # Gráfico ejecutivo simplificado
    if len(report_data) > 1 and numeric_columns:
        story.append(Paragraph("VISUALIZACIÓN EJECUTIVA", executive_subtitle_style))
        
        try:
            time_col = None
            for col in report_data[0].keys():
                if 'fecha' in col.lower() or 'date' in col.lower() or 'time' in col.lower():
                    time_col = col
                    break
            
            if time_col:
                # Crear gráfico ejecutivo
                fig, ax = plt.subplots(figsize=(8, 5))
                
                x_values = []
                y_values = []
                
                for row in report_data[:30]:  # Solo 30 puntos para gráfico ejecutivo
                    try:
                        time_val = row.get(time_col, '')
                        if isinstance(time_val, str):
                            try:
                                parsed_time = datetime.strptime(time_val, '%Y-%m-%d %H:%M:%S')
                                x_values.append(parsed_time)
                            except:
                                x_values.append(len(x_values))
                        else:
                            x_values.append(len(x_values))
                        
                        for num_col in numeric_columns[:1]:
                            val = row.get(num_col, 0)
                            if isinstance(val, str):
                                val = float(val) if val.replace('.', '').replace('-', '').isdigit() else 0
                            y_values.append(val if isinstance(val, (int, float)) else 0)
                            break
                    except:
                        continue
                
                if len(x_values) > 1 and len(y_values) > 1:
                    ax.plot(x_values, y_values, linewidth=3, color='#1e40af', alpha=0.8)
                    ax.fill_between(x_values, y_values, alpha=0.3, color='#1e40af')
                    ax.set_title(f'Tendencia de {numeric_columns[0]}', fontsize=14, fontweight='bold', color='#1e293b')
                    ax.set_xlabel('Tiempo', fontsize=11, color='#475569')
                    ax.set_ylabel(numeric_columns[0], fontsize=11, color='#475569')
                    ax.grid(True, alpha=0.2)
                    ax.spines['top'].set_visible(False)
                    ax.spines['right'].set_visible(False)
                    
                    if isinstance(x_values[0], datetime):
                        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
                        ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(x_values)//8)))
                        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, fontsize=9)
                    
                    plt.tight_layout()
                    
                    img_buffer = io.BytesIO()
                    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', 
                              facecolor='white', edgecolor='none')
                    img_buffer.seek(0)
                    
                    story.append(Image(img_buffer, width=10*cm, height=6*cm))
                    story.append(Spacer(1, 20))
                    
                    plt.close()
        except Exception as e:
            logger.warning(f"Error generando gráfico ejecutivo: {e}")
    
    # Conclusiones ejecutivas
    story.append(Paragraph("CONCLUSIONES EJECUTIVAS", executive_subtitle_style))
    story.append(Paragraph("• Los datos analizados muestran patrones consistentes en el período evaluado", executive_text_style))
    story.append(Paragraph("• Se identifican oportunidades de optimización basadas en las métricas clave", executive_text_style))
    story.append(Paragraph("• Se recomienda continuar el monitoreo para validar tendencias", executive_text_style))
    story.append(Spacer(1, 20))
    
    # Pie de página ejecutivo
    story.append(Paragraph("INFORMACIÓN DEL SISTEMA", executive_subtitle_style))
    story.append(Paragraph("• Sistema MTE - SIVE - Monitoreo Integral de Energía", executive_text_style))
    story.append(Paragraph("• Reporte generado automáticamente para análisis ejecutivo", executive_text_style))
    story.append(Paragraph("• Contacto: administrador@mtesive.com", executive_text_style))
    
    # Construir el PDF
    doc.build(story)
    
    # Calcular tamaño del archivo
    file_size = os.path.getsize(file_path)
    file_size_str = format_file_size(file_size)
    
    return file_size_str, len(report_data)


def format_file_size(size_bytes):
    """
    Formatea el tamaño del archivo en formato legible
    """
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    
    return f"{size_bytes:.1f} {size_names[i]}"


def get_report_status(task_id, user_id=None):
    """
    Obtiene el estado de un reporte.

    Si se pasa user_id, el reporte debe pertenecer a ese usuario (evita IDOR:
    que un usuario consulte el estado/URL de reportes de otro por task_id).
    """
    try:
        lookup = {'task_id': task_id}
        if user_id is not None:
            lookup['user_id'] = user_id
        report = GeneratedReport.objects.get(**lookup)
        
        status_info = {
            'task_id': report.task_id,
            'status': report.status,
            'progress': 0,
            'download_url': None,
            'error': None
        }
        
        if report.status == 'pending':
            status_info['progress'] = 0
        elif report.status == 'processing':
            status_info['progress'] = 50
        elif report.status == 'completed':
            status_info['progress'] = 100
            status_info['download_url'] = f"/api/reports/download/?task_id={task_id}"
        elif report.status == 'failed':
            status_info['progress'] = 0
            status_info['error'] = report.error_message
        
        return status_info
        
    except GeneratedReport.DoesNotExist:
        return None


def get_report_file(task_id, user_id=None):
    """
    Obtiene la información del archivo de un reporte.

    Si se pasa user_id, el reporte debe pertenecer a ese usuario (evita IDOR:
    que un usuario descargue reportes de otro conociendo el task_id).
    """
    try:
        lookup = {'task_id': task_id, 'status': 'completed'}
        if user_id is not None:
            lookup['user_id'] = user_id
        report = GeneratedReport.objects.get(**lookup)
        
        if not report.file_path or not os.path.exists(report.file_path):
            return None
        
        return {
            'file_path': report.file_path,
            'file_size': report.file_size,
            'record_count': report.record_count
        }
        
    except GeneratedReport.DoesNotExist:
        return None